#!/usr/bin/env python3
"""Импорт пакетов из Excel в формат core_packages.json.

Usage:
  python tools/import_core_packages_xlsx.py 3123131.xlsx --sheet "Лист1" --out frontend_shared/data/core_packages.json
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl  # type: ignore

RUB_PER_POINT = 4950

SEGMENT_MAP = {
    "Только розница": ("retail_only", "Только розница"),
    "Только опт": ("wholesale_only", "Только опт"),
    "Только производитель/импортер": ("producer_only", "Только производитель/импортер"),
    "Прозводитель розница": ("producer_retail", "Производитель + розница"),
    "Производитель розница": ("producer_retail", "Производитель + розница"),
    "Производитель + розница": ("producer_retail", "Производитель + розница"),
}


def _to_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _normalize_segment(value: str) -> Optional[Dict[str, str]]:
    key = str(value).strip()
    if key in SEGMENT_MAP:
        seg_key, title = SEGMENT_MAP[key]
        return {"segment_key": seg_key, "title": title}
    return None


def _ensure_group(groups: List[Dict[str, Any]], index: Dict[str, Dict[str, Any]], name: str) -> Dict[str, Any]:
    if name in index:
        return index[name]
    group = {
        "name": name,
        "points": 0,
        "note": "",
        "details": [],
    }
    index[name] = group
    groups.append(group)
    return group


def parse_xlsx(path: Path, sheet: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(path)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Лист {sheet!r} не найден. Доступно: {wb.sheetnames}")

    ws = wb[sheet]
    packages: List[Dict[str, Any]] = []

    current: Optional[Dict[str, Any]] = None
    groups: List[Dict[str, Any]] = []
    group_index: Dict[str, Dict[str, Any]] = {}

    for row in ws.iter_rows(values_only=True):
        a, b, c, d, e, f = (row + (None,) * 6)[:6]
        if a:
            seg = _normalize_segment(str(a))
            if seg:
                if current:
                    current["groups"] = groups
                    packages.append(current)
                current = {
                    "id": seg["segment_key"],
                    "title": seg["title"],
                    "segment_key": seg["segment_key"],
                    "total_points": None,
                    "price_rub": None,
                    "groups": [],
                }
                groups = []
                group_index = {}
                continue

        if not current:
            continue

        if a:
            group = _ensure_group(groups, group_index, str(a).strip())
            pts = _to_int(b)
            if pts is not None:
                group["points"] = pts
            group["note"] = str(c).strip() if c else ""

        if d and f:
            group = _ensure_group(groups, group_index, str(f).strip())
            detail = {"text": str(d).strip(), "points": _to_int(e) or 0}
            group["details"].append(detail)

        if not a and b and not d and not f:
            val = _to_int(b)
            if val is not None:
                if current.get("total_points") is None:
                    current["total_points"] = val
                elif current.get("price_rub") is None:
                    current["price_rub"] = val

    if current:
        current["groups"] = groups
        packages.append(current)

    for pkg in packages:
        total_points = pkg.get("total_points") or 0
        if not pkg.get("price_rub"):
            pkg["price_rub"] = total_points * RUB_PER_POINT

    return {"packages": packages}


def main() -> None:
    parser = argparse.ArgumentParser(description="Импорт пакетов из Excel в core_packages.json")
    parser.add_argument("xlsx", type=Path, help="Путь к .xlsx")
    parser.add_argument("--sheet", default="Лист1", help="Имя листа")
    parser.add_argument("--out", type=Path, default=Path("frontend_shared/data/core_packages.json"), help="Куда сохранить JSON")
    args = parser.parse_args()

    data = parse_xlsx(args.xlsx, args.sheet)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Сохранено: {args.out}")


if __name__ == "__main__":
    main()
